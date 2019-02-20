import openpyxl
import Getworkbook
import msvcrt


def conbinesheet(path):
    '''将文件目录多个温湿度记录表格提取平均值，并合并为一个表格'''
    a = Getworkbook.getFileName(path)
    print('以上文件已读取，按任意键继续')
    msvcrt.getch()
    wb2 = openpyxl.Workbook() #新建表格wb2
    ws2 = wb2.active #新建激活表单
    ws2.title = '每日数据'
    n = 1
    for worksheet in a:
        wb1 = openpyxl.load_workbook(worksheet)
        ws1 = wb1.active
        time_range = ws1['B4:B16']
        value_range = ws1['E4:E16']
        if worksheet[11:13] == '温度'or worksheet[12:14] == '温度':
            for i in range(n, n + 13):
                ws2.cell(row=i, column=1).value = time_range[i - n][0].value
                ws2.cell(row=i, column=2).value = worksheet[4:7]
                ws2.cell(row=i, column=3).value = value_range[i - n][0].value
        else:
            for i in range(n, n + 13):
                ws2.cell(row=i, column=4).value = value_range[i - n][0].value
            n += 13
    return wb2

if __name__ == '__main__':
    path = './'
    b = conbinesheet(path)
    b.active.insert_rows(0) # 插入表头行
    b.active['A1'] = '日期'
    b.active['B1'] = '机房'
    b.active['C1'] = '温度'
    b.active['D1'] = '湿度'
    b.save('summary.xlsx')
    print('数据汇总已保存在summary.xlsx')
    msvcrt.getch()
