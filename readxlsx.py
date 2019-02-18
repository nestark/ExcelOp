import openpyxl

wb = openpyxl.load_workbook("35F_34东-温湿度_湿度趋势分析_2019-02-01至2019-02-13.xlsx")
ws1 = wb.active
print(ws1[3][3].value)
print(type(wb))
print(wb.sheetnames)
print(ws1.title)
cell_range = ws1['A3':'E16']
tuple(ws1['A3':'E16'])
print(ws1['A3'].row + ws1['b2'].column)
cel = ws1['a3']
print(cel.coordinate)
print(type(cell_range))
for i in range(4, ws1.max_row,2):
    print(ws1.cell(i, 2).value)

for i in cell_range:
    for j in i:
        print(j.value)

